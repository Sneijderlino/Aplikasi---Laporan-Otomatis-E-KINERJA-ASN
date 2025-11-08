import logging
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

# Determine base directory (for resources) and database path under Local AppData
BASE_DIR = Path(__file__).resolve().parent
# Prefer the user's LOCALAPPDATA if available; fallback to <home>/AppData/Local
local_appdata = os.environ.get("LOCALAPPDATA")
if local_appdata:
    APPDATA_LOCAL = Path(local_appdata)
else:
    APPDATA_LOCAL = Path.home() / "AppData" / "Local"

# Folder specified by user: DataBase-Surat Masuk Keluar inside Local AppData
DB_DIR = APPDATA_LOCAL / "DataBase-Surat Masuk Keluar"
DB_FILE = DB_DIR / "DataBase.db"

# ================= Database ====================
class SuratDB:
    def __init__(self, db_file=DB_FILE):
        # Ensure DB parent directory exists before connecting
        try:
            db_path = Path(db_file)
            if not db_path.parent.exists():
                db_path.parent.mkdir(parents=True, exist_ok=True)
                logging.info("Created DB directory: %s", db_path.parent)
        except Exception:
            logging.exception("Failed to ensure DB directory exists")

        # ensure we pass a string path to sqlite and enable row factory for convenience
        logging.info("Opening database at %s", db_file)
        self.conn = sqlite3.connect(str(db_file))
        self.conn.row_factory = sqlite3.Row
        self.create_table()

    def create_table(self):
        sql = """
        CREATE TABLE IF NOT EXISTS surat (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jenis TEXT NOT NULL,
            nomor TEXT,
            tanggal TEXT,
            pihak TEXT,
            perihal TEXT,
            penanggung TEXT,
            catatan TEXT,
            created_at TEXT
        );
        """
        self.conn.execute(sql)
        self.conn.commit()

    def add_surat(self, jenis, nomor, tanggal, pihak, perihal, penanggung, catatan):
        sql = """
        INSERT INTO surat (jenis, nomor, tanggal, pihak, perihal, penanggung, catatan, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """
        self.conn.execute(sql, (jenis, nomor, tanggal, pihak, perihal, penanggung, catatan, datetime.datetime.now().isoformat()))
        self.conn.commit()

    def update_surat(self, surat_id, nomor, tanggal, pihak, perihal, penanggung, catatan):
        sql = """
        UPDATE surat SET nomor=?, tanggal=?, pihak=?, perihal=?, penanggung=?, catatan=? WHERE id=?
        """
        self.conn.execute(sql, (nomor, tanggal, pihak, perihal, penanggung, catatan, surat_id))
        self.conn.commit()

    def delete_surat(self, surat_id):
        self.conn.execute("DELETE FROM surat WHERE id=?", (surat_id,))
        self.conn.commit()

    def list_surat(self, jenis=None, search=""):
        sql = "SELECT id, nomor, tanggal, pihak, perihal, penanggung, catatan FROM surat WHERE jenis=?"
        params = [jenis]
        if search:
            sql += " AND (nomor LIKE ? OR pihak LIKE ? OR perihal LIKE ? OR penanggung LIKE ?)"
            s = f"%{search}%"
            params.extend([s, s, s, s])
        sql += " ORDER BY tanggal DESC, id DESC"
        cur = self.conn.execute(sql, params)
        return cur.fetchall()

    def close(self):
        self.conn.close()

# ================= GUI ====================
class SuratApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📄 Apk Pencatatan Surat Masuk Dan Keluar")
        self.root.geometry("1200x650")
        self.db = SuratDB()
        self.active_jenis = "masuk"
        self.selected_id = None
        self.logo_img = None
        self.logo_label = None
        self.active_button = None
        self.setup_styles()
        self.create_layout()
        self.load_surat()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        bg_color = "#1c1c2b" 
        fg_color = "#eaeaea"
        accent = "#4b8bd4"
        entry_bg = "#2b2b3a"
        self.root.configure(bg=bg_color)
        style.configure("TFrame", background=bg_color)
        style.configure("TLabel", background=bg_color, foreground=fg_color, font=("Segoe UI", 10))
        style.configure("TButton", background=accent, foreground="white", font=("Segoe UI", 10, "bold"), padding=5)
        style.map("TButton", background=[("active", "#357ab7")])
        style.configure("Clicked.TButton", background="#145A32", foreground="white", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview", background=entry_bg, fieldbackground=entry_bg, foreground=fg_color, font=("Segoe UI",10))
        style.configure("Treeview.Heading", background=bg_color, foreground=fg_color, font=("Segoe UI", 10, "bold"))
        style.map("Treeview", background=[('selected', accent)])

    def create_layout(self):
        bg_color = "#1c1c2b"
        # Header
        header = ttk.Frame(self.root, padding=(10, 10))
        header.pack(fill=tk.X)
        ttk.Label(header, text="📄 Apk Pencatatan Surat Masuk Dan Keluar", font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT)

        # Logo kantor (kanan atas)
        self.logo_label = ttk.Label(header)
        self.logo_label.pack(side=tk.RIGHT)
        self.load_default_logo()

        # Main Frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Sidebar
        sidebar = ttk.Frame(main_frame, width=180)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        ttk.Label(sidebar, text="Menu", font=("Segoe UI", 12, "bold")).pack(pady=(0, 10))

        self.btn_masuk = self.make_button(sidebar, "Surat Masuk", lambda: self.switch_tab("masuk", self.btn_masuk))
        self.btn_masuk.pack(fill=tk.X, pady=2)

        self.btn_keluar = self.make_button(sidebar, "Surat Keluar", lambda: self.switch_tab("keluar", self.btn_keluar))
        self.btn_keluar.pack(fill=tk.X, pady=2)

        # Content
        content = ttk.Frame(main_frame)
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Form Frame
        form_frame = ttk.LabelFrame(content, text="Input / Edit Surat")
        form_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(form_frame, text="Nomor:").grid(row=0, column=0, padx=5, pady=3, sticky=tk.W)
        self.ent_nomor = ttk.Entry(form_frame, width=30)
        self.ent_nomor.grid(row=0, column=1, padx=5, pady=3)

        ttk.Label(form_frame, text="Tanggal (DD-MM-YYYY):").grid(row=0, column=2, padx=5, pady=3, sticky=tk.W)
        self.ent_tanggal = ttk.Entry(form_frame, width=20)
        self.ent_tanggal.grid(row=0, column=3, padx=5, pady=3)
        self.ent_tanggal.insert(0, datetime.date.today().strftime("%d-%m-%Y"))

        ttk.Label(form_frame, text="Asal/Tujuan:").grid(row=1, column=0, padx=5, pady=3, sticky=tk.W)
        self.ent_pihak = ttk.Entry(form_frame, width=30)
        self.ent_pihak.grid(row=1, column=1, padx=5, pady=3)

        ttk.Label(form_frame, text="Perihal:").grid(row=1, column=2, padx=5, pady=3, sticky=tk.W)
        self.ent_perihal = ttk.Entry(form_frame, width=30)
        self.ent_perihal.grid(row=1, column=3, padx=5, pady=3)

        ttk.Label(form_frame, text="Penanggung Jawab:").grid(row=2, column=0, padx=5, pady=3, sticky=tk.W)
        self.ent_penanggung = ttk.Entry(form_frame, width=30)
        self.ent_penanggung.grid(row=2, column=1, padx=5, pady=3)

        ttk.Label(form_frame, text="Catatan:").grid(row=2, column=2, padx=5, pady=3, sticky=tk.W)
        self.ent_catatan = ttk.Entry(form_frame, width=30)
        self.ent_catatan.grid(row=2, column=3, padx=5, pady=3)

        # Button Frame
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=3, column=0, columnspan=4, pady=5)
        self.add_btn = self.make_button(btn_frame, "Tambah / Update", self.add_update_surat)
        self.add_btn.pack(side=tk.LEFT, padx=5)
        self.del_btn = self.make_button(btn_frame, "Hapus Terpilih", self.delete_selected)
        self.del_btn.pack(side=tk.LEFT, padx=5)
        self.exp_btn = self.make_button(btn_frame, "Ekspor Excel", self.export_excel)
        self.exp_btn.pack(side=tk.LEFT, padx=5)
        self.imp_btn = self.make_button(btn_frame, "Import Excel", self.import_excel)
        self.imp_btn.pack(side=tk.LEFT, padx=5)

        # Search
        search_frame = ttk.Frame(content)
        search_frame.pack(fill=tk.X, padx=5, pady=(0, 5))
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        self.ent_search = ttk.Entry(search_frame, textvariable=self.search_var)
        self.ent_search.pack(side=tk.LEFT, padx=5)
        self.make_button(search_frame, "Cari", self.search_surat).pack(side=tk.LEFT, padx=5)
        self.make_button(search_frame, "Reset", self.reset_search).pack(side=tk.LEFT, padx=5)

        # Table
        table_frame = ttk.Frame(content)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        columns = ("id", "nomor", "tanggal", "pihak", "perihal", "penanggung", "catatan")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            self.tree.heading(col, text=col.capitalize())
        widths = [50, 120, 100, 150, 200, 130, 250]
        for col, w in zip(columns, widths):
            self.tree.column(col, width=w)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", self.prefill_form)

        # Set awal tab aktif
        self.switch_tab("masuk", self.btn_masuk)

    # ================= Helper UI ====================
    def make_button(self, parent, text, cmd):
        btn = ttk.Button(parent, text=text)
        def on_click():
            self.set_active_button(btn)
            cmd()
        btn.configure(command=on_click)
        return btn

    def set_active_button(self, btn):
        if self.active_button and self.active_button != btn:
            self.active_button.configure(style="TButton")
        btn.configure(style="Clicked.TButton")
        self.active_button = btn
        self.root.bind("<Button-1>", self.reset_button_style_if_outside)

    def reset_button_style_if_outside(self, event):
        if not self.active_button:
            return
        widget = event.widget

        # Jangan reset warna jika tombol aktif adalah Surat Masuk atau Surat Keluar
        if self.active_button in [self.btn_masuk, self.btn_keluar]:
            return

        if widget != self.active_button:
            self.active_button.configure(style="TButton")
            self.active_button = None
            self.root.unbind("<Button-1>")
# Logo Aplikasi
    def load_default_logo(self):
        # Try a few likely locations for the logo image
        try:
            possible = [BASE_DIR / "logo_kantor.jpg", BASE_DIR / "icon" / "logo_kantor.jpg", Path("logo_kantor.jpg")]
            img_path = None
            for p in possible:
                if p and p.exists():
                    img_path = p
                    break
            if not img_path:
                logging.debug("No logo image found in expected locations")
                return
            img = Image.open(img_path)
            img = img.resize((100, 100), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(img)
            self.logo_label.configure(image=self.logo_img)
        except Exception as e:
            logging.exception("Failed to load logo image: %s", e)

    # ================= Actions ====================
    def switch_tab(self, jenis, btn):
        self.active_jenis = jenis
        self.selected_id = None
        self.clear_form()
        self.load_surat()
        self.btn_masuk.configure(style="TButton")
        self.btn_keluar.configure(style="TButton")
        btn.configure(style="Clicked.TButton")

    def on_close(self):
        """Cleanup before exit: close DB and destroy root."""
        try:
            if hasattr(self, "db") and self.db:
                self.db.close()
                logging.info("Database connection closed")
        except Exception:
            logging.exception("Error while closing database")
        try:
            self.root.destroy()
        except Exception:
            pass

    def add_update_surat(self):
        nomor = self.ent_nomor.get().strip()
        tanggal = self.ent_tanggal.get().strip()
        pihak = self.ent_pihak.get().strip()
        perihal = self.ent_perihal.get().strip()
        penanggung = self.ent_penanggung.get().strip()
        catatan = self.ent_catatan.get().strip()
        try:
            datetime.datetime.strptime(tanggal, "%d-%m-%Y")
        except:
            messagebox.showerror("Error", "Format tanggal salah! Gunakan DD-MM-YYYY")
            return
        if self.selected_id:
            self.db.update_surat(self.selected_id, nomor, tanggal, pihak, perihal, penanggung, catatan)
            messagebox.showinfo("Sukses", "Surat berhasil diupdate")
            self.selected_id = None
        else:
            self.db.add_surat(self.active_jenis, nomor, tanggal, pihak, perihal, penanggung, catatan)
            messagebox.showinfo("Sukses", "Surat berhasil ditambahkan")
        self.clear_form()
        self.load_surat()

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Pilih surat", "Pilih surat untuk dihapus")
            return
        surat_id = self.tree.item(sel[0], "values")[0]
        self.db.delete_surat(surat_id)
        self.load_surat()

    def load_surat(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        search = self.search_var.get().strip()
        rows = self.db.list_surat(self.active_jenis, search)
        for r in rows:
            self.tree.insert("", tk.END, values=r)

    def clear_form(self):
        self.ent_nomor.delete(0, tk.END)
        self.ent_tanggal.delete(0, tk.END)
        self.ent_tanggal.insert(0, datetime.date.today().strftime("%d-%m-%Y"))
        self.ent_pihak.delete(0, tk.END)
        self.ent_perihal.delete(0, tk.END)
        self.ent_penanggung.delete(0, tk.END)
        self.ent_catatan.delete(0, tk.END)
        self.selected_id = None

    def prefill_form(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        self.selected_id = vals[0]
        self.ent_nomor.delete(0, tk.END)
        self.ent_nomor.insert(0, vals[1])
        self.ent_tanggal.delete(0, tk.END)
        self.ent_tanggal.insert(0, vals[2])
        self.ent_pihak.delete(0, tk.END)
        self.ent_pihak.insert(0, vals[3])
        self.ent_perihal.delete(0, tk.END)
        self.ent_perihal.insert(0, vals[4])
        self.ent_penanggung.delete(0, tk.END)
        self.ent_penanggung.insert(0, vals[5])
        self.ent_catatan.delete(0, tk.END)
        self.ent_catatan.insert(0, vals[6])

    def export_excel(self):
        rows = [self.tree.item(i, "values") for i in self.tree.get_children()]
        if not rows:
            messagebox.showinfo("Tidak ada data", "Tidak ada surat untuk diekspor")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Surat"
        ws.append(["ID", "Nomor", "Tanggal", "Pihak", "Perihal", "Penanggung", "Catatan"])
        for r in rows:
            ws.append(r)
        try:
            wb.save(file_path)
            messagebox.showinfo("Sukses", f"Data berhasil diekspor ke {file_path}")
        except Exception as e:
            logging.exception("Failed to save Excel: %s", e)
            messagebox.showerror("Error", f"Gagal menyimpan Excel: {e}")

    def import_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    continue
                # validate row length and skip empty rows
                if not row or len(row) < 7:
                    logging.debug("Skipping invalid/short row at index %s: %s", idx, row)
                    continue
                # ignore ID column when present
                try:
                    if len(row) == 7:
                        nomor, tanggal, pihak, perihal, penanggung, catatan = row[1:7]
                    else:
                        # row may contain ID + 6 fields or more; try to pick expected columns
                        nomor = row[1]
                        tanggal = row[2]
                        pihak = row[3]
                        perihal = row[4]
                        penanggung = row[5]
                        catatan = row[6]
                except Exception:
                    logging.exception("Error unpacking row %s: %s", idx, row)
                    continue
                self.db.add_surat(self.active_jenis, nomor, tanggal, pihak, perihal, penanggung, catatan)
            self.load_surat()
            messagebox.showinfo("Sukses", "Data berhasil diimport dari Excel")
        except Exception as e:
            logging.exception("Failed to import Excel: %s", e)
            messagebox.showerror("Error", f"Gagal import Excel: {e}")

    def search_surat(self):
        self.load_surat()

    def reset_search(self):
        self.search_var.set("")
        self.load_surat()

# ================= Run ====================
def main():
    root = tk.Tk()
    app = SuratApp(root)
    # Ensure we close DB on window close
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()

if __name__ == "__main__":
    main()