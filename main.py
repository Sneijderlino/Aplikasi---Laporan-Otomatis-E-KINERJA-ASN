import os
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from dateutil.relativedelta import relativedelta # Menggunakan relativedelta untuk perhitungan bulan yang lebih aman
import pandas as pd
from pandas import Series

# Cek ketersediaan library opsional
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False
    Image = None
    ImageTk = None

try:
    from fpdf import FPDF
except Exception:
    FPDF = None

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ---------------- CONFIG ----------------

# !!! JALUR PENYIMPANAN ANDA - SESUAIKAN JIKA PERLU !!!
# Pastikan folder ini sudah ada atau bisa dibuat oleh aplikasi.
BASE_PATH = "D:\SNEIJDERLINO\DATABASE APLIKASI SNEIJDERLINO\DATABASE_E-KINERJA"

EXCEL_FILE = os.path.join(BASE_PATH, "laporan_asn.xlsx")
IDENT_FILE = os.path.join(BASE_PATH, "identitas_asn.json")

COLUMNS = [
    "Nama",
    "NIP",
    "Jabatan",
    "Unit Kerja",
    "Tanggal",
    "Waktu",
    "Uraian Kejadian",
    "Waktu Kebakaran",
    "Kerusakan",
    "Tindakan",
    "Foto",
    "Generated At"
]

# ----------------- Utility Functions ------------------------

def parse_date_flexible(date_str):
    """Mencoba parse tanggal dari berbagai format."""
    if not date_str:
        return None
    s = str(date_str).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    try:
        # Coba format ISO yang lebih fleksibel
        return datetime.fromisoformat(s.split()[0]).date()
    except Exception:
        return None

def to_ddmmyyyy(date_or_str):
    """Mengubah input tanggal menjadi format DD-MM-YYYY."""
    d = parse_date_flexible(date_or_str)
    return d.strftime("%d-%m-%Y") if d else ""

def ensure_base_dir(file_path):
    """Memastikan direktori tempat file akan disimpan itu ada."""
    dir_name = os.path.dirname(file_path)
    if dir_name and not os.path.exists(dir_name):
        try:
            os.makedirs(dir_name)
        except Exception as e:
            messagebox.showerror("Error Folder", f"Gagal membuat folder tujuan: {dir_name}\n{e}")
            raise # Menghentikan eksekusi jika gagal membuat folder

def get_previous_month_date(date_obj=None):
    """
    Mendapatkan objek datetime.date untuk bulan sebelumnya.
    Menggunakan relativedelta untuk penanganan bulan yang lebih aman.
    """
    if date_obj is None:
        date_obj = datetime.now()
    
    # Hitung 1 bulan sebelumnya
    prev_month_dt = date_obj - relativedelta(months=1)
    return prev_month_dt.date()

# ----------------- Excel helpers ------------------------

def ensure_excel():
    """Membuat file Excel jika belum ada."""
    if not os.path.exists(EXCEL_FILE):
        df = pd.DataFrame(columns=COLUMNS)
        df.to_excel(EXCEL_FILE, index=False, engine="openpyxl")

def load_data():
    """Memuat data dari file Excel."""
    try:
        ensure_base_dir(EXCEL_FILE) 
    except Exception:
        return pd.DataFrame(columns=COLUMNS) 
        
    ensure_excel()
    try:
        df = pd.read_excel(EXCEL_FILE, engine="openpyxl")

        for c in COLUMNS:
            if c not in df.columns:
                df[c] = ""

        return df.reindex(columns=COLUMNS).copy()
    except Exception as e:
        # Jika file ada tapi rusak atau kosong, kembalikan DataFrame kosong
        if 'Worksheet named' in str(e): # Handle corrupted/empty sheet
             return pd.DataFrame(columns=COLUMNS)
        messagebox.showerror("Error", f"Gagal membaca data: {e}")
        return pd.DataFrame(columns=COLUMNS)

def save_data(df):
    """Menyimpan DataFrame ke file Excel."""
    try:
        ensure_base_dir(EXCEL_FILE) 
        df = df.reindex(columns=COLUMNS)
        df.to_excel(EXCEL_FILE, index=False, engine="openpyxl")
    except Exception as e:
        messagebox.showerror("Error", f"Gagal menyimpan data: {e}")

def _insert_image_into_excel_last_row(foto_path):
    """Menyisipkan gambar ke kolom 'Foto' di baris terakhir Excel."""
    if not (PIL_AVAILABLE and OPENPYXL_AVAILABLE):
        return
    if not foto_path or not os.path.exists(foto_path):
        return
    try:
        ensure_base_dir(EXCEL_FILE)
        
        # NOTE: Ini akan mengoverwrite file Excel, jadi pastikan dilakukan setelah save_data
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if "Foto" not in headers:
            wb.close()
            return
        
        target_row = ws.max_row
        if target_row < 2: 
            wb.close()
            return

        col_idx = headers.index("Foto") + 1
        
        img = Image.open(foto_path)
        max_w = 160 # Lebar maksimum di Excel
        w, h = img.size
        
        if w > max_w:
            ratio = max_w / float(w)
            try:
                img = img.resize((max_w, int(h * ratio)), Image.Resampling.LANCZOS)
            except AttributeError:
                img = img.resize((max_w, int(h * ratio)), Image.LANCZOS)
        
        # Simpan file sementara (PNG lebih universal untuk openpyxl)
        tmp_path = os.path.join(os.path.dirname(EXCEL_FILE), "__tmp_img.png")
        img.save(tmp_path)

        xi = XLImage(tmp_path)
        xi.anchor = f"{get_column_letter(col_idx)}{target_row}"
        ws.add_image(xi)
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        try:
            os.remove(tmp_path)
        except Exception:
            pass
            
    except Exception as e:
        print("Warning: gagal sisip gambar ke Excel:", e) 

# ----------------- Utility Functions (Styling) ------------------------

def style_button_hover(btn, normal_bg=None, hover_bg=None, active_bg=None):
    """Menambahkan efek hover ke tombol tk."""
    try:
        if isinstance(btn, ttk.Button):
            return # Skip ttk buttons, use style map instead
    except Exception:
        pass
    
    try:
        current = btn.cget("bg")
    except Exception:
        current = normal_bg or "#f0f0f0"
        
    normal = normal_bg or current
    hover = hover_bg or normal
    active = active_bg or hover
    
    try:
        btn.config(bg=normal, activebackground=active, cursor="hand2")
    except Exception:
        pass
        
    def on_enter(e):
        try:
            btn['bg'] = hover
        except Exception:
            pass
            
    def on_leave(e):
        try:
            btn['bg'] = normal
        except Exception:
            pass
            
    def on_press(e):
        try:
            btn['bg'] = active
        except Exception:
            pass
            
    def on_release(e):
        try:
            btn['bg'] = hover
        except Exception:
            pass
            
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    btn.bind("<ButtonPress-1>", on_press)
    btn.bind("<ButtonRelease-1>", on_release) 

# ----------------- identity persistence ------------------------

def load_identitas():
    """Memuat identitas ASN yang tersimpan dari JSON."""
    try:
        ensure_base_dir(IDENT_FILE)
    except Exception:
        return {}
        
    if os.path.exists(IDENT_FILE):
        try:
            with open(IDENT_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_identitas(data):
    """Menyimpan identitas ASN ke JSON."""
    try:
        ensure_base_dir(IDENT_FILE) 
        with open(IDENT_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("Gagal menyimpan identitas:", e)

def hapus_identitas_file():
    """Menghapus file identitas ASN."""
    try:
        if os.path.exists(IDENT_FILE):
            os.remove(IDENT_FILE)
    except Exception as e:
        print("Gagal hapus identitas:", e)

# ----------------- Main App Class ------------------------

class LaporanApp:
    def __init__(self, root):
        self.root = root
        root.title("Aplikasi Ekinerja — Laporan Kejadian Kebakaran")
        root.geometry("1280x800") # Ukuran awal yang sedikit lebih pendek
        self.selected_foto_path = ""
        self.preview_img = None
        self.initial_preview_size = (380, 320)
        
        # --- KONFIGURASI ROOT UNTUK RESPONSIF ---
        root.grid_rowconfigure(3, weight=1)    
        root.grid_columnconfigure(0, weight=1) 
        
        # --- TEMA/STYLING PROFESIONAL --------
        style = ttk.Style()
        style.theme_use('clam')
        
        # WARNA
        self.PRIMARY_COLOR = "#5D4AA0"
        self.ACCENT_COLOR = "#785BB8"
        self.LIGHT_BG = "#F4F4F9"
        self.DARK_TEXT = "#333333"
        self.WHITE_TEXT = "white"
        
        root.configure(bg=self.LIGHT_BG)

        # Style Header
        style.configure("Header.TFrame", background=self.PRIMARY_COLOR)
        style.configure("Header.TLabel", background=self.PRIMARY_COLOR, foreground=self.WHITE_TEXT, font=("Segoe UI", 16, "bold"))
        style.configure("SubHeader.TLabel", background=self.PRIMARY_COLOR, foreground=self.WHITE_TEXT, font=("Segoe UI", 12))
        
        # Style Frame Input
        style.configure("Form.TFrame", background=self.LIGHT_BG)
        style.configure("FormLabel.TLabel", background=self.LIGHT_BG, foreground=self.DARK_TEXT)
        
        # Style Tombol
        style.configure("TButton", font=("Segoe UI", 10), padding=5, background="#E0E0E0", foreground=self.DARK_TEXT)
        style.map("TButton", background=[('active', '#C0C0C0')])
        
        # Style Tombol Aksi Utama (CRUD & Export)
        style.configure("Add.TButton", background="#4CAF50", foreground=self.WHITE_TEXT)
        style.map("Add.TButton", background=[('active', '#66BB6A')])
        style.configure("Edit.TButton", background="#2196F3", foreground=self.WHITE_TEXT)
        style.map("Edit.TButton", background=[('active', '#42A5F5')])
        style.configure("Delete.TButton", background="#F44336", foreground=self.WHITE_TEXT)
        style.map("Delete.TButton", background=[('active', '#E57373')])
        style.configure("Reset.TButton", background="#FF9800", foreground=self.WHITE_TEXT)
        style.map("Reset.TButton", background=[('active', '#FFB74D')])
        style.configure("Upload.TButton", background=self.ACCENT_COLOR, foreground=self.WHITE_TEXT)
        style.map("Upload.TButton", background=[('active', self.PRIMARY_COLOR)])
        
        # Style Treeview (Tabel)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background=self.ACCENT_COLOR, foreground=self.WHITE_TEXT)
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)
        style.map("Treeview", background=[('selected', self.ACCENT_COLOR)], foreground=[('selected', self.WHITE_TEXT)])

        # Header (Row 0)
        header = ttk.Frame(root, style="Header.TFrame", height=52)
        header.grid(row=0, column=0, sticky="ew")
        ttk.Label(header, text="APLIKASI BY SNEIJDERLINO ", style="Header.TLabel").pack(side="left", padx=12, pady=4)
        ttk.Label(header, text="Laporan koordinasi dengan kepala regu terkait informasi kejadian kebakaran", style="SubHeader.TLabel").pack(side="left", padx=12)
        btn_logout = tk.Button(header, text="Log Out", bg="#DCD7FF", relief="flat", command=self.on_logout)
        btn_logout.pack(side="right", padx=12, pady=4)
        style_button_hover(btn_logout, normal_bg="#DCD7FF", hover_bg="#E9E6FF", active_bg="#C0B5FF")

        # Form top (Kontainer Utama Formulir) - Row 1
        form_container = ttk.Frame(root, style="Form.TFrame", padding="10 5 10 5")
        form_container.grid(row=1, column=0, sticky="ew")
        
        # Frame yang menampung 3 kolom: Left, Mid, Right
        form_frame = ttk.Frame(form_container, style="Form.TFrame")
        form_frame.pack(fill="x", expand=True) 
        
        # Konfigurasi Bobot Kolom pada form_frame
        form_frame.grid_columnconfigure(0, weight=0) 
        form_frame.grid_columnconfigure(1, weight=3) 
        form_frame.grid_columnconfigure(2, weight=1) 
        form_frame.grid_rowconfigure(0, weight=1)    

        # ------------------- Left basic fields (Kolom 0) -------------------
        left = ttk.Frame(form_frame, style="Form.TFrame", padding="6 4 12 4")
        left.grid(row=0, column=0, sticky="nsw")
        
        self.entries = {}
        labels = ["Nama", "NIP", "Jabatan", "Unit Kerja", "Tanggal", "Waktu"]
        
        # Field Entry
        for i, field in enumerate(labels):
            lbl = ttk.Label(left, text=field + ":", style="FormLabel.TLabel", anchor="w")
            lbl.grid(row=i, column=0, sticky="w", pady=2, padx=(0, 8)) 
            ent = ttk.Entry(left, width=35)
            ent.grid(row=i, column=1, pady=2, sticky="ew") 
            self.entries[field] = ent
        
        # --- PENAMBAHAN LABEL BULAN (1 BULAN SEBELUMNYA) ---
        
        # Hitung tanggal 1 bulan sebelumnya
        prev_month_date = get_previous_month_date()
        prev_month_name = prev_month_date.strftime("%B %Y")
        
        # Simpan bulan/tahun untuk digunakan di judul PDF
        self.ekinerja_month_year = prev_month_name.upper()
        
        tanggal_row = labels.index("Tanggal") 
        waktu_row = labels.index("Waktu")     

        # Geser Label dan Entri Tanggal/Waktu ke bawah
        self.entries["Tanggal"].grid(row=tanggal_row + 1, column=1, pady=2, sticky="ew")
        left.grid_slaves(row=tanggal_row, column=0)[0].grid(row=tanggal_row + 1) 
        
        self.entries["Waktu"].grid(row=waktu_row + 1, column=1, pady=2, sticky="ew")
        left.grid_slaves(row=waktu_row, column=0)[0].grid(row=waktu_row + 1) 
        
        # Sisipkan Label Bulan
        lbl_bulan = ttk.Label(left, 
                              text=f"Bulan Ekinerja: {self.ekinerja_month_year} 📅", 
                              style="FormLabel.TLabel", 
                              font=("Segoe UI", 10, "bold"), 
                              foreground=self.ACCENT_COLOR, 
                              anchor="w")
        lbl_bulan.grid(row=tanggal_row, column=0, columnspan=2, sticky="ew", pady=(8, 4), padx=(0,0)) 
        
        left.grid_columnconfigure(1, weight=1) 
        
        self.entries["Tanggal"].delete(0, "end")
        self.entries["Tanggal"].insert(0, datetime.now().strftime("%d-%m-%Y"))
        self.entries["Waktu"].delete(0, "end")
        self.entries["Waktu"].insert(0, datetime.now().strftime("%H:%M:%S"))
        
        self._load_and_lock_identitas()
        self.entries["Nama"].bind("<Double-1>", self._on_nama_double_click)
        self._create_identitas_context_menu()
        
        # ------------------- Middle multiline (Kolom 1) -------------------
        mid = ttk.Frame(form_frame, style="Form.TFrame", padding="6 4")
        mid.grid(row=0, column=1, sticky="nsew", padx=15) 
        self.texts = {}
        
        mid.grid_columnconfigure(0, weight=1) 
        multiline = ["Uraian Kejadian", "Waktu Kebakaran", "Kerusakan", "Tindakan"]
        
        for i, field in enumerate(multiline):
            lbl = ttk.Label(mid, text=field + ":", style="FormLabel.TLabel", anchor="w")
            lbl.grid(row=i*2, column=0, sticky="w", pady=(2,0)) 
            
            txt_frame = ttk.Frame(mid)
            txt_frame.grid(row=i*2+1, column=0, pady=(1, 5), sticky="ew") 
            
            txt = tk.Text(txt_frame, width=1, height=4, wrap="word", relief="flat", borderwidth=1, highlightthickness=1, highlightcolor=self.PRIMARY_COLOR, highlightbackground="#CCCCCC")
            txt.pack(side="left", fill="both", expand=True) 
            
            vscroll = ttk.Scrollbar(txt_frame, orient="vertical", command=txt.yview)
            vscroll.pack(side="right", fill="y")
            txt.config(yscrollcommand=vscroll.set)
            self.texts[field] = txt

        mid.grid_rowconfigure(len(multiline)*2 - 1, weight=1) 
        
        # ------------------- Right preview (Kolom 2) -------------------
        right = ttk.Frame(form_frame, style="Form.TFrame", padding="8 4")
        right.grid(row=0, column=2, sticky="nsew", padx=(20, 0))
        right.grid_rowconfigure(0, weight=1) 
        right.grid_columnconfigure(0, weight=1) 

        preview_container = ttk.LabelFrame(right, text="Preview Foto", padding="10 10 10 10")
        preview_container.grid(row=0, column=0, padx=6, pady=4, sticky="nsew") 
        
        self.foto_label = tk.Label(preview_container, text="[Area Preview Foto]", width=1, height=1, relief="flat", bg="#EEEEEE", fg="#888888", anchor="center", font=("Segoe UI", 10))
        self.foto_label.pack(fill="both", expand=True)
        self.foto_label.bind("<Configure>", self._on_preview_resize)
        
        btn_frame_foto = ttk.Frame(right, padding="0 8 0 0")
        btn_frame_foto.grid(row=1, column=0, sticky="ew")
        
        btn_upload = ttk.Button(btn_frame_foto, text="📸 Upload Foto", command=self.pilih_foto, style="Upload.TButton")
        btn_upload.pack(side="left", padx=(0, 10))
        btn_hapus_foto = ttk.Button(btn_frame_foto, text="🗑 Hapus Foto", command=self.reset_foto)
        btn_hapus_foto.pack(side="left")

        # =========================================================================
        # Actions & Filter Container - Row 2
        # =========================================================================

        action_filter_container = ttk.Frame(root, padding="12 4")
        action_filter_container.grid(row=2, column=0, sticky="ew")
        
        action_filter_container.grid_rowconfigure(0, weight=0) 
        action_filter_container.grid_rowconfigure(1, weight=0) 
        action_filter_container.grid_columnconfigure(0, weight=1)

        # ------------------- Action Buttons (Row 0 di kontainer) -------------------
        actions = ttk.Frame(action_filter_container)
        actions.grid(row=0, column=0, sticky="ew", pady=(0, 6)) 

        # Container untuk Tombol CRUD (Kiri)
        crud_frame = ttk.Frame(actions)
        crud_frame.pack(side="left") 
        
        btn_tambah = ttk.Button(crud_frame, text="Tambah Laporan", width=18, command=self.tambah_laporan, style="Add.TButton")
        btn_tambah.pack(side="left", padx=3)
        btn_edit = ttk.Button(crud_frame, text="Update Laporan", width=18, command=self.edit_laporan, style="Edit.TButton")
        btn_edit.pack(side="left", padx=3)
        btn_hapus = ttk.Button(crud_frame, text="Hapus Laporan", width=18, command=self.hapus_laporan, style="Delete.TButton")
        btn_hapus.pack(side="left", padx=3)
        btn_reset = ttk.Button(crud_frame, text="Reset Form", width=18, command=self.reset_form, style="Reset.TButton")
        btn_reset.pack(side="left", padx=6)

        # Container untuk Tombol EXPORT (Kanan)
        export_frame = ttk.Frame(actions)
        export_frame.pack(side="right") 

        btn_export_excel = ttk.Button(export_frame, text="Export Semua → Excel", command=self.export_all_excel, style="Edit.TButton")
        btn_export_excel.pack(side="right", padx=3) 
        btn_export_many = ttk.Button(export_frame, text="Export Semua → PDF (Banyak)", command=self.export_all_multiple_pdfs, style="Reset.TButton")
        btn_export_many.pack(side="right", padx=3)
        btn_export_all_pdf = ttk.Button(export_frame, text="Export Semua → PDF (Satu)", command=self.export_all_single_pdf, style="Add.TButton")
        btn_export_all_pdf.pack(side="right", padx=3)
        btn_export_sel = ttk.Button(export_frame, text="Export Pilih → PDF", command=self.export_selected_pdf, style="Edit.TButton")
        btn_export_sel.pack(side="right", padx=3) 

        # ------------------- Filter Box (Row 1 di kontainer) -------------------
        filter_frame = ttk.LabelFrame(action_filter_container, text="Filter Data", padding="4 2 4 2") 
        filter_frame.grid(row=1, column=0, sticky="ew") 

        
        # Kolom Filter
        ttk.Label(filter_frame, text="Kolom:", style="FormLabel.TLabel").grid(row=0, column=0, padx=2, sticky="w")
        self.filter_col = ttk.Combobox(filter_frame, values=["Nama", "NIP", "Jabatan", "Unit Kerja", "Uraian Kejadian", "Waktu Kebakaran", "Kerusakan", "Tindakan"], width=13, state="readonly") 
        self.filter_col.set("Nama")
        self.filter_col.grid(row=0, column=1, padx=2, pady=1)
        
        # Keyword Filter
        ttk.Label(filter_frame, text="Keyword:", style="FormLabel.TLabel").grid(row=0, column=2, padx=4, sticky="w")
        self.filter_kw = ttk.Entry(filter_frame, width=13) 
        self.filter_kw.grid(row=0, column=3, padx=2, pady=1, sticky="ew") 
        
        # Date From Filter
        ttk.Label(filter_frame, text="Tgl From:", style="FormLabel.TLabel").grid(row=0, column=4, padx=6, sticky="w") 
        self.date_from = ttk.Entry(filter_frame, width=8) 
        self.date_from.grid(row=0, column=5, padx=2, pady=1, sticky="ew")
        
        # Date To Filter
        ttk.Label(filter_frame, text="To:", style="FormLabel.TLabel").grid(row=0, column=6, padx=4, sticky="w")
        self.date_to = ttk.Entry(filter_frame, width=8) 
        self.date_to.grid(row=0, column=7, padx=2, pady=1, sticky="ew")
        
        # Tombol Filter
        btn_apply_filter = ttk.Button(filter_frame, text="🔍 Filter", command=self.apply_filter, style="Edit.TButton")
        btn_apply_filter.grid(row=0, column=8, padx=4)
        
        # Tombol Reset
        btn_reset_filter = ttk.Button(filter_frame, text="🔄 Reset", command=self.reset_filter, style="Reset.TButton")
        btn_reset_filter.grid(row=0, column=9, padx=2)
        
        # Kolom Keyword, From, dan To diberi bobot
        filter_frame.grid_columnconfigure(3, weight=1)  
        filter_frame.grid_columnconfigure(5, weight=1)
        filter_frame.grid_columnconfigure(7, weight=1)


        # Table area (MENGEMBANG PENUH) - Row 3
        table_frame = ttk.Frame(root, padding="12 4 12 8")
        table_frame.grid(row=3, column=0, sticky="nsew") 
        
        cols = COLUMNS.copy()
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        
        # Konfigurasi Lebar Kolom Tabel (Optimal)
        for c in cols:
            self.tree.heading(c, text=c)
            if c in ("Uraian Kejadian","Waktu Kebakaran","Kerusakan","Tindakan"):
                self.tree.column(c, width=120, anchor="w") 
            elif c in ("Nama", "Jabatan", "Unit Kerja"):
                self.tree.column(c, width=100, anchor="w")
            elif c in ("Tanggal", "Waktu"):
                self.tree.column(c, width=80, anchor="center")
            elif c=="Foto":
                self.tree.column(c, width=80, anchor="center")
            elif c=="Generated At":
                self.tree.column(c, width=120, anchor="center")
            else:
                self.tree.column(c, width=100, anchor="w")
                
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Penempatan Treeview dan Scrollbar
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Konfigurasi Bobot pada table_frame
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        self.tree.bind("<Double-1>", self.on_row_double_click)
        self.tampilkan_data()

    # ---------------- PDF helpers (DIJADIKAN METHOD) ------------------------

    def _write_report_to_pdf_page(self, pdf: "FPDF", row: pd.Series):
        """Menulis detail satu baris laporan ke halaman PDF dengan tata letak rapi."""
        
        # --- 1. HEADER & JUDUL LAPORAN ---
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 5, "Dinas Satuan Polisi Pamong Praja Dan Pemadam Kebakaran", ln=True, align="C")
        pdf.cell(0, 5, "Kabupaten Maluku Barat Daya", ln=True, align="C")
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 5, "Kota Tiakur", ln=True, align="C")
        pdf.ln(3)
        
        # GARIS PEMISAH SETELAH KOP
        pdf.line(pdf.get_x(), pdf.get_y(), 200, pdf.get_y()) 
        pdf.ln(5)

        pdf.set_font("Arial", "B", 14)
        # JUDUL TELAH DIBUAT DINAMIS BERDASARKAN BULAN SEBELUMNYA
        pdf.cell(0, 8, f"LAPORAN E-KINERJA BULAN {self.ekinerja_month_year}", ln=True, align="C")
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 5, "Koordinasi dengan Kepala Regu terkait informasi kejadian kebakaran", ln=True, align="C")
        
        pdf.ln(5) 

        # --- 2. DATA IDENTITAS (Menggunakan Tabel) ---
        pdf.set_font("Arial", "", 10)
        
        # Data Pasangan Kunci-Nilai untuk Identitas
        identitas_data = [
            ("Nama", str(row.get("Nama", "") or "")),
            ("NIP", str(row.get("NIP", "") or "")),
            ("Jabatan", str(row.get("Jabatan", "") or "")),
            ("Unit Kerja", str(row.get("Unit Kerja", "") or "")),
            ("Tanggal", to_ddmmyyyy(row.get("Tanggal", ""))),
            ("Waktu", str(row.get("Waktu", "") or "")),
        ]

        col_width_label = 25
        col_width_sep = 5

        for label, value in identitas_data:
            x = pdf.get_x()
            y = pdf.get_y()
            
            pdf.set_font("Arial", "B", 10)
            pdf.cell(col_width_label, 5, f"{label}", 0) # Label
            
            pdf.set_xy(x + col_width_label, y)
            pdf.set_font("Arial", "", 10)
            pdf.cell(col_width_sep, 5, ":", 0) # Pemisah
            
            pdf.set_xy(x + col_width_label + col_width_sep, y)
            pdf.cell(0, 5, value, 0, 1) # Nilai (ln=1)

        pdf.ln(5)

        # --- 3. DETAIL KEJADIAN ---
        detail_fields = ["Uraian Kejadian", "Waktu Kebakaran", "Kerusakan", "Tindakan"]
        
        for field in detail_fields:
            pdf.set_font("Arial", "B", 11)
            # Judul dengan garis bawah (border='B')
            pdf.cell(0, 7, f"{field}:", ln=True, border='B') 
            pdf.set_font("Arial", "", 10)
            
            text = str(row.get(field, "") or "")
            # Gunakan multi_cell dengan align='J' (Justify) agar rata kiri-kanan
            pdf.multi_cell(0, 5, text, 0, 'J') 
            pdf.ln(4)
            
        # --- 4. FOTO BUKTI ---
        pdf.set_font("Arial", "B", 11)
        # Judul dengan garis bawah (border='B')
        pdf.cell(0, 7, "Foto Bukti:", ln=True, border='B')
        pdf.ln(3)

        foto_path = row.get("Foto", "")
        
        if pd.notna(foto_path) and foto_path and os.path.exists(foto_path):
            try:
                # Pusatkan gambar dan batasi lebar (w=120, x=45)
                pdf.image(foto_path, w=120, x=45) 
                pdf.ln(4)
            except Exception:
                pdf.set_font("Arial", "I", 10)
                pdf.cell(0, 5, "[Gagal menampilkan foto. Pastikan format JPG/PNG]", ln=True)
        else:
            pdf.set_font("Arial", "I", 10)
            pdf.cell(0, 5, "[Tidak ada foto]", ln=True)
            
        # --- 5. FOOTER (Generated At) ---
        pdf.ln(6)
        pdf.set_font("Arial", "I", 8)
        pdf.cell(0, 5, f"Dibuat: {row.get('Generated At', '')}", ln=True, align="R") 

    def export_all_to_single_pdf(self, df, out_path):
        """Export semua data ke satu file PDF."""
        if FPDF is None:
            messagebox.showwarning("Missing lib", "Install 'fpdf' (pip install fpdf) untuk export PDF.")
            return
        
        pdf = FPDF()
        for _, row in df.iterrows():
            pdf.add_page()
            self._write_report_to_pdf_page(pdf, row)
            
        try:
            pdf.output(out_path)
            messagebox.showinfo("Sukses", f"PDF semua laporan dibuat:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membuat PDF: {e}")

    def export_each_row_to_pdf_files(self, df, out_dir):
        """Export setiap baris data ke file PDF terpisah."""
        if FPDF is None:
            messagebox.showwarning("Missing lib", "Install 'fpdf' (pip install fpdf) untuk export PDF.")
            return
        
        os.makedirs(out_dir, exist_ok=True)
        created = []
        
        for idx, row in df.iterrows():
            # Buat nama file yang aman dan deskriptif
            nama_safe = "".join(c for c in str(row.get("Nama","")) if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_") or f"row{idx}"
            tanggal_safe = str(to_ddmmyyyy(row.get("Tanggal",""))).replace("-", "")
            
            filename = os.path.join(out_dir, f"laporan_{nama_safe}_{tanggal_safe}_{idx+1}.pdf")
            
            pdf = FPDF()
            pdf.add_page()
            self._write_report_to_pdf_page(pdf, row)
            
            try:
                pdf.output(filename)
                created.append(filename)
            except Exception:
                pass
                
        if created:
            messagebox.showinfo("Sukses", f"{len(created)} file PDF dibuat di:\n{out_dir}")
        else:
            messagebox.showwarning("Hasil", "Tidak ada file PDF berhasil dibuat.") 

    # ---------------- identity functions ----------------
    
    def _load_and_lock_identitas(self):
        """Memuat identitas tersimpan dan mengunci field."""
        data = load_identitas()
        for k in ["Nama","NIP","Jabatan","Unit Kerja"]:
            ent = self.entries.get(k)
            if ent:
                try:
                    ent.config(state="normal")
                    ent.delete(0, "end")
                except Exception:
                    pass
                
                if data and k in data:
                    ent.insert(0, data.get(k, ""))
                    try:
                        ent.config(state="readonly")
                    except Exception:
                        pass

    def _create_identitas_context_menu(self):
        """Membuat menu konteks untuk field Nama."""
        self.ident_menu = tk.Menu(self.root, tearoff=0)
        self.ident_menu.add_command(label="Hapus identitas tersimpan", command=self._on_hapus_identitas)
        self.entries["Nama"].bind("<Button-3>", lambda e: self.ident_menu.post(e.x_root, e.y_root))

    def _on_nama_double_click(self, event=None):
        """Mengunci/membuka field identitas dan menyimpan data."""
        ent = self.entries["Nama"]
        state = ent.cget("state")
        
        if state == "readonly":
            # Buka kunci untuk edit
            for k in ["Nama","NIP","Jabatan","Unit Kerja"]:
                try:
                    self.entries[k].config(state="normal")
                except Exception:
                    pass
            messagebox.showinfo("Edit Identitas", "Field identitas dibuka. Ubah data lalu double-click lagi di Nama untuk menyimpan.")
            return
        else:
            # Simpan dan kunci
            data = {}
            for k in ["Nama","NIP","Jabatan","Unit Kerja"]:
                data[k] = self.entries[k].get().strip()
            
            if not data.get("Nama"):
                messagebox.showwarning("Validasi", "Nama tidak boleh kosong saat menyimpan identitas.")
                return
            
            nip_val = data.get("NIP","")
            if nip_val and not self._validate_nip(nip_val):
                messagebox.showwarning("Validasi", "NIP harus berupa angka dan minimal 5 digit.")
                return
            
            save_identitas(data)
            
            for k in ["Nama","NIP","Jabatan","Unit Kerja"]:
                try:
                    self.entries[k].config(state="readonly")
                except Exception:
                    pass
                    
            messagebox.showinfo("Sukses", "Identitas ASN disimpan dan dikunci. (Double-click Nama untuk ubah lagi)")

    def _on_hapus_identitas(self):
        """Menghapus data identitas dari file."""
        if not os.path.exists(IDENT_FILE):
            messagebox.showinfo("Info", "Tidak ada identitas tersimpan.")
            return
            
        if not messagebox.askyesno("Konfirmasi", "Yakin ingin menghapus identitas ASN yang tersimpan?"):
            return
            
        hapus_identitas_file()
        
        # Buka kunci field dan kosongkan isinya
        for k in ["Nama","NIP","Jabatan","Unit Kerja"]:
            try:
                self.entries[k].config(state="normal")
                self.entries[k].delete(0, "end")
            except Exception:
                pass
                
        messagebox.showinfo("Sukses", "Identitas ASN berhasil dihapus.")
        
    def _save_new_identitas_if_needed(self):
        """Menawarkan untuk menyimpan identitas jika belum tersimpan/terkunci."""
        if self.entries.get("Nama") and self.entries["Nama"].cget("state") == "readonly":
            return 
            
        data = {}
        for k in ["Nama", "NIP", "Jabatan", "Unit Kerja"]:
            data[k] = self.entries[k].get().strip()
        
        if data.get("Nama") and data.get("NIP") and self._validate_nip(data.get("NIP")):
            if messagebox.askyesno("Simpan Identitas?", 
                                   "Apakah Anda ingin **menyimpan** Nama, NIP, Jabatan, dan Unit Kerja ini agar terisi otomatis dan terkunci di laporan berikutnya?"):
                save_identitas(data)

                for k in ["Nama", "NIP", "Jabatan", "Unit Kerja"]:
                    try:
                        self.entries[k].config(state="readonly")
                    except Exception:
                        pass

    # ---------------- photo handlers ----------------
    
    def pilih_foto(self):
        """Membuka dialog untuk memilih file foto."""
        if not PIL_AVAILABLE:
            messagebox.showwarning("Missing lib", "Install 'Pillow' (pip install pillow) untuk upload & preview foto.")
            return
            
        path = filedialog.askopenfilename(title="Pilih foto", filetypes=[("Image Files", "*.jpg *.jpeg *.png *.bmp")])
        if path:
            self.selected_foto_path = path
            self._render_preview_image(path)

    def _on_preview_resize(self, event):
        """Menangani resize preview foto saat jendela diubah ukurannya."""
        if self.selected_foto_path:
            self._render_preview_image(self.selected_foto_path)

    def _render_preview_image(self, path):
        """Menampilkan gambar di area preview dengan skala yang tepat."""
        if not PIL_AVAILABLE:
            return
        
        try:
            img = Image.open(path)
            lbl = self.foto_label
            
            # Ambil ukuran label saat ini
            w = lbl.winfo_width()
            h = lbl.winfo_height()
            
            # Gunakan ukuran default jika widget belum dirender atau ukurannya kecil
            if w < 50 or h < 50:
                w, h = self.initial_preview_size
                
            iw, ih = img.size
            
            # Hitung rasio untuk fit di dalam label
            ratio = min(w / iw, h / ih)
            if ratio <= 0:
                ratio = 1.0
            
            new_size = (max(1, int(iw * ratio)), max(1, int(ih * ratio)))
            
            try:
                img_resized = img.resize(new_size, Image.Resampling.LANCZOS)
            except AttributeError:
                img_resized = img.resize(new_size, Image.LANCZOS)
                
            photo = ImageTk.PhotoImage(img_resized)
            lbl.configure(image=photo, text="")
            lbl.image = photo # Jaga referensi
            self.selected_foto_path = path
            
        except Exception as e:
            lbl = self.foto_label
            lbl.configure(image="", text="[Preview gagal atau format tidak didukung]", fg="red")
            print("Preview error:", e)

    def reset_foto(self):
        """Menghapus foto dari form dan preview."""
        self.selected_foto_path = ""
        self.foto_label.configure(image="", text="[Area Preview Foto]", bg="#EEEEEE", fg="#888888")
        try:
            # Hapus referensi agar gambar bisa di-garbage collected
            del self.foto_label.image 
        except Exception:
            pass

    # ---------------- form helpers ----------------
    
    def get_form_data(self):
        """Mengambil semua data dari formulir input."""
        data = {}
        
        for k, ent in self.entries.items():
            val = ent.get().strip()
            
            if k == "Tanggal":
                # Pastikan tanggal disimpan dalam format konsisten (DD-MM-YYYY)
                if val:
                    parsed = parse_date_flexible(val)
                    if parsed:
                        val = parsed.strftime("%d-%m-%Y")
                    else:
                        # Jika tidak bisa diparse, gunakan tanggal hari ini
                        val = datetime.now().strftime("%d-%m-%Y") 
                
            data[k] = val
            
        texts_map = {
            "Uraian Kejadian": self.texts["Uraian Kejadian"],
            "Waktu Kebakaran": self.texts["Waktu Kebakaran"],
            "Kerusakan": self.texts["Kerusakan"],
            "Tindakan": self.texts["Tindakan"],
        }
        
        for k, txt in texts_map.items():
            data[k] = txt.get("1.0", "end").strip()
            
        data["Foto"] = self.selected_foto_path or ""
        data["Generated At"] = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        
        return data

    def set_form_data_from_row(self, row):
        """Mengisi formulir input dengan data dari baris DataFrame."""
        # 1. Handle Entry fields
        for k in ["Nama","NIP","Jabatan","Unit Kerja","Tanggal","Waktu"]:
            ent = self.entries.get(k)
            if not ent: continue
            
            val = row.get(k,"")
            if pd.isna(val): val = ""
                
            if k == "Tanggal" and val:
                # Tampilkan dalam format DD-MM-YYYY
                parsed = parse_date_flexible(val)
                if parsed:
                    val = parsed.strftime("%d-%m-%Y")
            
            current_state = ent.cget("state")
            
            # Tangani field yang readonly (Identitas)
            if current_state == "readonly":
                ent.config(state="normal")
                ent.delete(0, "end")
                ent.insert(0, str(val))
                ent.config(state="readonly")
            else:
                ent.delete(0, "end")
                ent.insert(0, str(val))

        # 2. Handle Text fields (Multiline)
        texts_map = {
            "Uraian Kejadian": self.texts["Uraian Kejadian"],
            "Waktu Kebakaran": self.texts["Waktu Kebakaran"],
            "Kerusakan": self.texts["Kerusakan"],
            "Tindakan": self.texts["Tindakan"],
        }
        
        for k in texts_map.keys():
            txt = self.texts.get(k)
            if not txt: continue
            
            txt.delete("1.0", "end")
            v = row.get(k,"")
            if pd.isna(v): v = ""
            txt.insert("1.0", str(v))

        # 3. Handle Foto
        foto = row.get("Foto","")
        if pd.notna(foto) and foto and os.path.exists(foto):
            self.selected_foto_path = foto
            if PIL_AVAILABLE:
                self._render_preview_image(foto)
            else:
                self.foto_label.configure(text="[Foto ada, instal Pillow untuk preview]")
        else:
            self.reset_foto()

    def tampilkan_data(self, filtered_df=None):
        """Memuat dan menampilkan data ke Treeview."""
        for r in self.tree.get_children():
            self.tree.delete(r)
            
        df = filtered_df if filtered_df is not None else load_data()
        
        for idx, row in df.reset_index(drop=True).iterrows():
            vals = []
            for c in COLUMNS:
                v = row.get(c,"")
                if pd.isna(v): v = ""
                
                if c == "Tanggal" and v:
                    # Pastikan tampilan di tabel DD-MM-YYYY
                    parsed = parse_date_flexible(v)
                    if parsed:
                        v = parsed.strftime("%d-%m-%Y")
                        
                vals.append(str(v))
                
            self.tree.insert("", "end", iid=str(idx), values=vals)

    # ---------------- Validations ----------------
    
    def _validate_date(self, s):
        """Memvalidasi format tanggal."""
        return parse_date_flexible(s) is not None

    def _validate_time(self, s):
        """Memvalidasi format waktu (HH:MM:SS atau HH:MM)."""
        if not s: 
            return False
        for fmt in ("%H:%M:%S","%H:%M"):
            try:
                datetime.strptime(s, fmt)
                return True
            except Exception:
                continue
        return False

    def _validate_nip(self, nip):
        """Memvalidasi NIP (minimal 5 digit angka)."""
        n = str(nip).strip()
        return n.isdigit() and len(n) >= 5 

    # ---------------- Preview Pop-up ----------------
    
    def _show_data_preview_popup(self, data, action_text, confirm_callback):
        """Menampilkan pop-up preview data sebelum disimpan/diupdate."""
        # Membuat jendela pop-up
        preview_window = tk.Toplevel(self.root)
        preview_window.title(f"Konfirmasi Laporan: {action_text.upper()}")
        preview_window.geometry("850x550") 
        preview_window.transient(self.root) 
        preview_window.grab_set() 

        # Gaya untuk jendela pop-up
        style = ttk.Style()
        style.configure("Preview.TLabel", background=self.LIGHT_BG, foreground=self.DARK_TEXT)
        
        main_frame = ttk.Frame(preview_window, padding="10", style="Form.TFrame")
        main_frame.pack(fill="both", expand=True)

        # KONFIGURASI RESPONSIF DI POP-UP
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(2, weight=1) 

        # Header preview (Row 0)
        ttk.Label(main_frame, text=f"Periksa sebelum **{action_text.upper()}**:", style="Preview.TLabel", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 10))

        # Tampilan Data Ringkas (Row 1)
        preview_cols = ["Nama", "NIP", "Jabatan", "Unit Kerja", "Tanggal", "Waktu"]
        preview_tree = ttk.Treeview(main_frame, columns=preview_cols, show="headings", selectmode="none", height=1) 
        
        style.configure("PreviewTree.Heading", font=("Segoe UI", 10, "bold"), background="#333333", foreground="white") 
        preview_tree.tag_configure('oddrow', background='white')
        
        for col in preview_cols:
            preview_tree.heading(col, text=col.replace(" ", "\n"))
            if col in ("Tanggal", "Waktu"):
                preview_tree.column(col, width=80, anchor="center")
            elif col == "NIP":
                preview_tree.column(col, width=120, anchor="w")
            elif col == "Nama":
                preview_tree.column(col, width=150, anchor="w")
            else:
                preview_tree.column(col, width=120, anchor="w")
                
        preview_tree.grid(row=1, column=0, sticky="ew", padx=5, pady=5)

        preview_values = []
        for col in preview_cols:
            val = data.get(col, "")
            if col == "Tanggal":
                val = to_ddmmyyyy(val)
            preview_values.append(str(val))
            
        preview_tree.insert("", "end", values=preview_values, tags=('oddrow',))

        # Detail Multiline dan Foto (Row 2)
        detail_frame = ttk.Frame(main_frame, style="Form.TFrame", padding="5")
        detail_frame.grid(row=2, column=0, sticky="nsew", pady=(15, 0))
        detail_frame.grid_columnconfigure(1, weight=1) 
        
        fields_detail = [("Uraian Kejadian:", "Uraian Kejadian", 0, 4), 
                         ("Waktu Kebakaran:", "Waktu Kebakaran", 1, 2), 
                         ("Kerusakan:", "Kerusakan", 2, 4), 
                         ("Tindakan:", "Tindakan", 3, 4)]
        
        for i, (lbl_txt, data_key, row_idx, height) in enumerate(fields_detail):
            ttk.Label(detail_frame, text=lbl_txt, style="Preview.TLabel", font=("Segoe UI", 10, "bold")).grid(row=row_idx, column=0, sticky="nw", padx=(0, 10), pady=(5, 0))
            
            txt_container = ttk.Frame(detail_frame)
            txt_container.grid(row=row_idx, column=1, sticky="nsew", padx=(0, 15), pady=(5, 10))
            
            txt = tk.Text(txt_container, width=1, height=height, wrap="word", relief="flat", borderwidth=1, state="disabled")
            txt.pack(fill="both", expand=True) 
            
            txt.config(state="normal")
            txt.insert("1.0", data.get(data_key, f"Tidak ada {data_key.lower().replace(' ', '')} tercatat."))
            txt.config(state="disabled")
            
            # Beri bobot pada baris Tindakan agar mengembang
            if data_key == "Tindakan": 
                 detail_frame.grid_rowconfigure(row_idx, weight=1) 


        ttk.Label(detail_frame, text="Foto Bukti:", style="Preview.TLabel", font=("Segoe UI", 10, "bold")).grid(row=4, column=0, sticky="nw", pady=(10, 0))
        
        # Penempatan Foto
        foto_container = ttk.Frame(detail_frame)
        foto_container.grid(row=4, column=1, sticky="w", pady=(5, 0))
        detail_frame.grid_rowconfigure(4, weight=0) 

        foto_path = data.get("Foto", "")
        if PIL_AVAILABLE and foto_path and os.path.exists(foto_path):
            img = Image.open(foto_path)
            w, h = img.size
            max_w = 250
            ratio = min(max_w / float(w), 1.0) 
            
            try:
                img_resized = img.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            except AttributeError:
                img_resized = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
                
            photo = ImageTk.PhotoImage(img_resized)
            foto_label = tk.Label(foto_container, image=photo, relief="solid", borderwidth=1)
            foto_label.image = photo
            foto_label.pack()
        else:
            ttk.Label(foto_container, text="[Tidak ada foto / Gagal load]", style="Preview.TLabel", foreground="red").pack()

        # Tombol Aksi (Row 3)
        btn_frame = ttk.Frame(main_frame, style="Form.TFrame")
        btn_frame.grid(row=3, column=0, sticky="ew", pady=15)
        
        btn_batal = tk.Button(btn_frame, text="X Batal", command=preview_window.destroy, bg="#F44336", fg="white", font=("Segoe UI", 10, "bold"), padx=10, pady=5, relief="flat")
        btn_batal.pack(side="right", padx=10)
        style_button_hover(btn_batal, normal_bg="#F44336", hover_bg="#E57373")
        
        btn_konfirmasi = tk.Button(btn_frame, text="✔ Konfirmasi", command=lambda: [preview_window.destroy(), confirm_callback()], bg="#4CAF50", fg="white", font=("Segoe UI", 10, "bold"), padx=10, pady=5, relief="flat")
        btn_konfirmasi.pack(side="right")
        style_button_hover(btn_konfirmasi, normal_bg="#4CAF50", hover_bg="#66BB6A")
        
        preview_window.wait_window(preview_window)

    # ---------------- CRUD operations ----------------
    
    def tambah_laporan(self):
        """Menambahkan laporan baru."""
        self._save_new_identitas_if_needed()
        data = self.get_form_data()

        # --- Validasi ---
        if not data.get("Nama"):
            messagebox.showwarning("Validasi", "Nama harus diisi.")
            return
        if not self._validate_nip(data.get("NIP","")):
            messagebox.showwarning("Validasi", "NIP harus berupa angka dan minimal 5 digit.")
            return
        if not self._validate_date(data.get("Tanggal","")):
            messagebox.showwarning("Format tanggal", "Format 'Tanggal' tidak dikenali. Gunakan DD-MM-YYYY.")
            return
        if not self._validate_time(data.get("Waktu","")):
            messagebox.showwarning("Format waktu", "Format 'Waktu' tidak dikenali. Gunakan HH:MM:SS.")
            return

        # --- Panggil Preview Pop-up ---
        self._show_data_preview_popup(data, "Simpan", lambda: self._perform_tambah_laporan(data))

    def _perform_tambah_laporan(self, data):
        """Logika penyimpanan data setelah konfirmasi."""
        # --- Proses Simpan ---
        df = load_data()
        row_to_add = {c: data.get(c, "") for c in COLUMNS}
        
        # Gunakan pd.concat yang lebih modern daripada df.append
        df = pd.concat([df, pd.DataFrame([row_to_add], columns=COLUMNS)], ignore_index=True)
        save_data(df)
        
        # Hanya sisipkan gambar ke Excel jika tersedia
        if data.get("Foto") and OPENPYXL_AVAILABLE:
            _insert_image_into_excel_last_row(data.get("Foto"))
            
        self.tampilkan_data()
        messagebox.showinfo("Sukses", "Laporan berhasil ditambahkan.")
        self.reset_form()

    def hapus_laporan(self):
        """Menghapus laporan yang dipilih."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Peringatan", "Pilih baris yang akan dihapus.")
            return

        if not messagebox.askyesno("Konfirmasi", "Yakin menghapus laporan?"):
            return

        idx = int(sel[0])
        df = load_data()
        
        if idx < 0 or idx >= len(df):
            messagebox.showerror("Error", "Indeks tidak valid.")
            return

        # Drop baris berdasarkan indeks yang dipilih
        df = df.drop(index=df.index[idx]).reset_index(drop=True)
        save_data(df)
        
        self.tampilkan_data()
        messagebox.showinfo("Sukses", "Laporan dihapus.")

    def edit_laporan(self):
        """Mengupdate laporan yang dipilih."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Peringatan", "Pilih baris yang akan diedit (double-click untuk load).")
            return
            
        idx = int(sel[0])
        df = load_data()
        
        if idx < 0 or idx >= len(df):
            messagebox.showerror("Error", "Indeks tidak valid.")
            return

        new = self.get_form_data()

        # Pertahankan waktu pembuatan asli
        original_generated_at = df.iloc[idx].get("Generated At", datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
        new["Generated At"] = original_generated_at

        # --- Validasi ---
        if not new.get("Nama"):
            messagebox.showwarning("Validasi", "Nama harus diisi.")
            return
        if not self._validate_nip(new.get("NIP","")):
            messagebox.showwarning("Validasi", "NIP harus berupa angka dan minimal 5 digit.")
            return
        if not self._validate_date(new.get("Tanggal","")):
            messagebox.showwarning("Format tanggal", "Format 'Tanggal' tidak dikenali. Gunakan DD-MM-YYYY.")
            return
        if not self._validate_time(new.get("Waktu","")):
            messagebox.showwarning("Format waktu", "Format 'Waktu' tidak dikenali. Gunakan HH:MM:SS.")
            return

        # --- Panggil Preview Pop-up ---
        self._show_data_preview_popup(new, "Update", lambda: self._perform_edit_laporan(idx, new))

    def _perform_edit_laporan(self, idx, new_data):
        """Logika update data setelah konfirmasi."""
        # --- Proses Update ---
        df = load_data()
        
        # Pastikan kita bekerja dengan indeks yang benar
        df_index_to_update = df.index[idx]
        
        for k, v in new_data.items():
            if k not in df.columns:
                df[k] = ""
            df.at[df_index_to_update, k] = v
            
        save_data(df)
        
        # Peringatan: Update foto di Excel setelah edit masih kompleks dan tidak bisa dilakukan secara langsung tanpa re-generate seluruh Excel.
        
        self.tampilkan_data()
        messagebox.showinfo("Sukses", "Laporan diperbarui.")

    def reset_form(self):
        """Mengosongkan semua field input (kecuali identitas yang terkunci)."""
        # Reset Entry fields
        for k, ent in self.entries.items():
            try:
                # Hanya reset jika field tidak dikunci (readonly)
                if ent.cget("state") != "readonly":
                    ent.delete(0, "end")
            except Exception:
                pass 

        # Isi kembali Tanggal dan Waktu default
        if self.entries["Tanggal"].cget("state") != "readonly":
            self.entries["Tanggal"].delete(0, "end")
            self.entries["Tanggal"].insert(0, datetime.now().strftime("%d-%m-%Y"))
        
        if self.entries["Waktu"].cget("state") != "readonly":
            self.entries["Waktu"].delete(0, "end")
            self.entries["Waktu"].insert(0, datetime.now().strftime("%H:%M:%S"))

        # Reset Text fields
        texts_map = {
            "Uraian Kejadian": self.texts["Uraian Kejadian"],
            "Waktu Kebakaran": self.texts["Waktu Kebakaran"],
            "Kerusakan": self.texts["Kerusakan"],
            "Tindakan": self.texts["Tindakan"],
        }
        for txt in texts_map.values():
            txt.delete("1.0", "end")
            
        self.reset_foto()

    def on_row_double_click(self, event):
        """Memuat data baris yang di-double-click ke formulir."""
        sel = self.tree.selection()
        if not sel:
            return
            
        idx = int(sel[0])
        df = load_data()
        
        if idx < 0 or idx >= len(df):
            return
            
        row = df.iloc[idx]
        self.set_form_data_from_row(row)

    # ---------------- Filter / Search ----------------
    
    def apply_filter(self):
        """Menerapkan filter keyword dan/atau tanggal."""
        df = load_data()
        col = self.filter_col.get().strip()
        kw = self.filter_kw.get().strip().lower()
        df_filtered = df.copy()

        # 1. Filter Keyword
        if kw:
            if col and col in df.columns:
                df_filtered = df_filtered[df_filtered[col].astype(str).str.lower().str.contains(kw, na=False)]
            else:
                # Cari di semua kolom yang relevan jika kolom tidak dipilih
                mask = False
                for c in ["Nama","NIP","Jabatan","Unit Kerja","Uraian Kejadian","Waktu Kebakaran","Kerusakan","Tindakan"]:
                    if c in df_filtered.columns:
                        mask = mask | df_filtered[c].astype(str).str.lower().str.contains(kw, na=False)
                df_filtered = df_filtered[mask]
                
        df_filtered2 = df_filtered.copy()

        # 2. Filter Tanggal
        dfrom = self.date_from.get().strip()
        dto = self.date_to.get().strip()
        dfrom_dt = None
        dto_dt = None

        if dfrom:
            try:
                dfrom_dt = parse_date_flexible(dfrom)
                if not dfrom_dt: raise ValueError("Tanggal From tidak valid")
            except Exception:
                messagebox.showwarning("Format tanggal", "Format 'Tanggal From' harus DD-MM-YYYY atau YYYY-MM-DD")
                return
                
        if dto:
            try:
                dto_dt = parse_date_flexible(dto)
                if not dto_dt: raise ValueError("Tanggal To tidak valid")
            except Exception:
                messagebox.showwarning("Format tanggal", "Format 'Tanggal To' harus DD-MM-YYYY atau YYYY-MM-DD")
                return

        if dfrom_dt or dto_dt:
            def _date_in_range_filter(date_str):
                parsed = parse_date_flexible(date_str)
                if not parsed: return False
                
                in_range = True
                if dfrom_dt:
                    in_range = in_range and (parsed >= dfrom_dt)
                if dto_dt:
                    in_range = in_range and (parsed <= dto_dt)
                return in_range
                
            df_filtered2 = df_filtered2[df_filtered2["Tanggal"].astype(str).apply(_date_in_range_filter)]

        self.tampilkan_data(df_filtered2.reset_index(drop=True))

    def reset_filter(self):
        """Mengatur ulang semua filter dan menampilkan semua data."""
        self.filter_kw.delete(0, "end")
        self.date_from.delete(0, "end")
        self.date_to.delete(0, "end")
        self.filter_col.set("Nama")
        self.tampilkan_data()

    # ---------------- Exports ----------------

    def export_selected_pdf(self):
        """Export baris yang dipilih ke satu file PDF."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Peringatan", "Pilih baris yang akan di-export.")
            return

        df = load_data().reset_index(drop=True)
        idx = int(sel[0])
        
        if idx < 0 or idx >= len(df):
            messagebox.showerror("Error", "Indeks tidak valid.")
            return
            
        row = df.iloc[idx]
        
        nama_safe = "".join(c for c in str(row.get("Nama","")) if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_") or "laporan_selected"
        tanggal_safe = str(to_ddmmyyyy(row.get("Tanggal",""))).replace("-", "")
        
        out_path = filedialog.asksaveasfilename(title="Simpan PDF", defaultextension=".pdf", filetypes=[("PDF","*.pdf")], initialfile=f"laporan_{nama_safe}_{tanggal_safe}.pdf")
        
        if out_path:
            if FPDF is None:
                messagebox.showwarning("Missing lib", "Install 'fpdf' (pip install fpdf) untuk export PDF.")
                return
                
            pdf = FPDF()
            pdf.add_page()
            self._write_report_to_pdf_page(pdf, row)
            
            try:
                pdf.output(out_path)
                messagebox.showinfo("Sukses", f"PDF dibuat: {out_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal menyimpan PDF: {e}")

    def export_all_single_pdf(self):
        """Export semua data ke satu file PDF tunggal."""
        df = load_data()
        if df.empty:
            messagebox.showwarning("Kosong", "Tidak ada data untuk di-export.")
            return
            
        out_path = filedialog.asksaveasfilename(title="Simpan semua laporan ke 1 PDF", defaultextension=".pdf", filetypes=[("PDF","*.pdf")], initialfile=f"laporan_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        
        if not out_path: 
            return
            
        self.export_all_to_single_pdf(df, out_path)

    def export_all_multiple_pdfs(self):
        """Export semua data ke file PDF terpisah per baris."""
        df = load_data()
        if df.empty:
            messagebox.showwarning("Kosong", "Tidak ada data untuk di-export.")
            return
            
        out_dir = filedialog.askdirectory(title="Pilih folder tujuan (per-file PDFs)")
        
        if not out_dir: 
            return
            
        self.export_each_row_to_pdf_files(df, out_dir)

    def export_all_excel(self):
        """Export semua data ke file Excel."""
        df = load_data()
        if df.empty:
            messagebox.showwarning("Kosong", "Tidak ada data untuk di-export.")
            return
            
        out_path = filedialog.asksaveasfilename(title="Simpan semua data ke Excel", defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile=f"laporan_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        if not out_path: 
            return
            
        df_export = df.copy()
        
        if "Tanggal" in df_export.columns:
            # Pastikan format tanggal di kolom Excel adalah DD-MM-YYYY untuk kemudahan baca
            df_export["Tanggal"] = df_export["Tanggal"].apply(lambda x: to_ddmmyyyy(x) if x else "")
            
        try:
            df_export.to_excel(out_path, index=False, engine="openpyxl")
            messagebox.showinfo("Sukses", f"Excel dibuat: {out_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan Excel: {e}\nPastikan library 'openpyxl' terinstal.")

    # ---------------- Logout ----------------

    def on_logout(self):
        """Keluar dari aplikasi."""
        if messagebox.askyesno("Logout", "Yakin ingin logout?"):
            try:
                self.root.destroy()
            except Exception:
                os._exit(0) 

# ---------------- main ----------------

def main():
    # Cek dan beri pesan jika ada library penting yang kurang
    if not (PIL_AVAILABLE and OPENPYXL_AVAILABLE and FPDF):
        missing = []
        if not PIL_AVAILABLE: missing.append("Pillow (pip install pillow) untuk gambar")
        if not OPENPYXL_AVAILABLE: missing.append("openpyxl (pip install openpyxl) untuk Excel")
        # NOTE: Library 'python-dateutil' diperlukan untuk get_previous_month_date, 
        # tapi seringkali sudah ada. Jika tidak, tambahkan pesan:
        # missing.append("python-dateutil (pip install python-dateutil) untuk tanggal")
        if not FPDF: missing.append("fpdf (pip install fpdf) untuk PDF")
        
        if missing:
            print(f"PERINGATAN: Library berikut tidak terinstal dan fungsionalitas terkait akan dinonaktifkan: {', '.join(missing)}")
            
    root = tk.Tk()
    try:
        style = ttk.Style(root)
        style.theme_use('clam')
    except Exception:
        pass
        
    app = LaporanApp(root)
    root.mainloop() 
    
if __name__ == "__main__":
    main()